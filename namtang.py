import discord
import openpyxl
import request
import bs4
import asyncio
import datetime

client = discord.Client()


@client.event
async def on_ready():
    print("남탕대장 출근!")
    print(client.user.name)
    print(client.user.id)
    print("===========")
    print("이것은 조선관리인 봇의 클라이언트입니다")
    print("본 클라이언트의 저작권은 세콤에게 있습니다")
    print("현재 남탕관리인이 사용할수 있는 명령어는")
    print("!명령어,!변태순위 입니다")
    print("===========")

    await client.change_presence(game=discord.Game(name="위대한 수령동지를 위하여 | !도움말", type=1))


@client.event
async def on_member_join(member):
    role = ""
    for i in member.server.roles:
        if i.name == "린민":
            role = i
            break
    await client.add_roles(member, role)
    fmt = '{0.mention} 동무가 월북했디 모두 축하해주라우!'
    channel = member.server.get_channel("625685144258478083")
    await client.send_message(channel, fmt.format(member, member.server))

@client.event
async def on_member_remove(member):
    channel = member.server.get_channel("625685144258478083")
    fmt = '{0.mention} 동무가 탈북했다우 간나새끼..'
    await client.send_message(channel, fmt.format(member, member.server))

@client.event
async def on_message(message):
    if message.author.bot:
        return None

    id = message.author.id
    channel = message.channel

    if message.content.startswith("!임시공지"):
        channel = message.channel
        embed = discord.Embed(
            title='임시공지',
            description='임시공지가 없습네다',
            colour=discord.Colour.blue()
        )

        embed.set_footer(text='이상이라우')
        embed.add_field(name='제목', value='임시공지가 없습네다', inline=False)
        embed.set_image(url="https://img3.yna.co.kr/photo/yna/YH/2011/12/23/PYH2011122304620001300_P2.jpg")

        await client.send_message(channel, embed=embed)

    if message.content.startswith("!도움말"):
        channel = message.channel
        embed = discord.Embed(
            title='도움말',
            description='내래 도움말이오',
            colour=discord.Colour.blue()
        )

        embed.set_footer(text='이상이라우')
        embed.add_field(name='!학습', value='내래 이몸을 가르치라우 !학습 [호출단어] [말할단어]', inline=False)
        embed.add_field(name='!기억', value='내래 가르친걸 기억한다우 !기억 [호출단어]', inline=False)
        embed.add_field(name='!역할설정', value='내래 역할을 준다우 !역할설정 [ID] [역할이름]', inline=False)
        embed.add_field(name='!하는일', value='내래 밥값을 하는지 보여준다우', inline=False)
        embed.add_field(name='!정보', value='인-포마이숀을 보여주겠디', inline=False)
        embed.add_field(name='!경고', value='경고를 주겠디 !경고 [ID]', inline=False)
        embed.add_field(name='!확인', value='경고 횟수를 확인해주겠디 !확인 [ID]', inline=False)
        embed.add_field(name='!움짤', value='움짤모음을 알려주겠디', inline=False)
        embed.add_field(name='!첩보요원', value='가입일, 닉네임등의 정보를 알려주겠디', inline=False)
        embed.set_image(url="https://img3.yna.co.kr/photo/yna/YH/2011/12/23/PYH2011122304620001300_P2.jpg")

        await client.send_message(channel, embed=embed)

    if message.content.startswith("!움짤"):
        channel = message.channel
        embed = discord.Embed(
            title='움짤들',
            colour=discord.Colour.blue()
        )

        embed.set_footer(text='이상이라우')
        embed.add_field(name='!굿타임', value='좋은시간되세요 뺑뺑이', inline=False)
        embed.add_field(name='!반작용', value='자살골짤', inline=False)
        embed.add_field(name='!놀람', value='존나 놀라는짤', inline=False)
        embed.add_field(name='!잉아쌀라말라이꿈', value='노코멘트', inline=False)

        await client.send_message(channel, embed=embed)

    if message.content.startswith("!첩보요원"):
        date = datetime.datetime.utcfromtimestamp(((int(message.author.id) >> 22) + 1420070400000) / 1000)
        now = datetime.datetime.now()
        nowDate = now.strftime('%Y-%m-%d')
        embed = discord.Embed(color=0x00ff00)
        embed.add_field(name="이름", value=message.author.name, inline=True)
        embed.add_field(name="서버닉네임", value=message.author.display_name, inline=True)
        embed.add_field(name="가입일", value=str(date.year) + "년" + str(date.month) + "월" + str(date.day) + "일" , inline=True)
        embed.add_field(name="오늘자", value=nowDate , inline=True)
        embed.add_field(name="아이디", value=message.author.id, inline=True)
        embed.set_thumbnail(url=message.author.avatar_url)
        await client.send_message(message.channel, embed=embed)

    if message.content.startswith("!굿타임"):
        channel = message.channel
        embed = discord.Embed(
            title='',
            description='',
            colour=discord.Colour.blue()
        )

        embed.set_image(url="https://cdn.discordapp.com/emojis/579875968802619418.gif?v=1")

        await client.send_message(channel, embed=embed)

    if message.content.startswith("!반작용"):
        channel = message.channel
        embed = discord.Embed(
            title='',
            description='',
            colour=discord.Colour.blue()
        )

        embed.set_image(url="https://cdn.discordapp.com/attachments/591252897430241284/591256228102012931/2eb30b699416ab7faec36216c123de16_UGwyBUUa3JR4ICi7TXOOU.gif")

        await client.send_message(channel, embed=embed)

    if message.content.startswith("!놀람"):
        channel = message.channel
        embed = discord.Embed(
            title='',
            description='',
            colour=discord.Colour.blue()
        )

        embed.set_image(url="https://cdn.discordapp.com/attachments/591252897430241284/591256228651466771/C0E7B9F2_1.gif")

        await client.send_message(channel, embed=embed)

    if message.content.startswith("!잉아쌀라말라이꿈"):
        channel = message.channel
        embed = discord.Embed(
            title='',
            description='',
            colour=discord.Colour.blue()
        )

        embed.set_image(url="https://cdn.discordapp.com/attachments/591252897430241284/591256618633789442/205b613e7255f18dc94966ac66ff1f94.gif")

        await client.send_message(channel, embed=embed)

    if message.content.startswith("!정보"):
        channel = message.channel
        embed = discord.Embed(
            title='인-포마이숀',
            description='',
            colour=discord.Colour.blue()
        )

        embed.set_footer(text='이상이라우')
        embed.add_field(name='제작',
                        value='제작자 : 시바 존나멋진 김민규(ㅈ한민국 기준 16세)\n 마지막 패치 : 갱신귀찮아서 안씀 ㅅㄱ\n 제작 이유 : 위대한 조선 민주주의 노휘빈 공화국의 입주자를 받기 위해서',
                        inline=False)
        embed.add_field(name='제작된 날', value='서기 2019년 6월 15일 11시경', inline=False)
        embed.add_field(name='고위간부', value='수령 : 노휘빈\n군 : 김애용\n 원로 : 김민규\n 친족:대현', inline=False)
        embed.add_field(name='!하는일', value='내래 밥값을 하는지 보여준다우', inline=False)
        embed.set_image(url="https://img3.yna.co.kr/photo/yna/YH/2011/12/23/PYH2011122304620001300_P2.jpg")

        await client.send_message(channel, embed=embed)

    if message.content.startswith("!하는일"):
        channel = message.channel
        embed = discord.Embed(
            title='하는일',
            description='내래 밥값한다우',
            colour=discord.Colour.blue()
        )

        embed.set_footer(text='이상이라우')
        embed.add_field(name='린민 지급', value='신입들에게 린민역할을 지급한다우', inline=False)
        embed.add_field(name='봇', value='이건 모르겠다우 시바꺼', inline=False)
        embed.add_field(name='월북알림', value='월북, 탈북 정보를 알려준다우', inline=False)
        embed.set_image(url="https://img3.yna.co.kr/photo/yna/YH/2011/12/23/PYH2011122304620001300_P2.jpg")

        await client.send_message(channel, embed=embed)


    if message.content.startswith("!학습"):
        file = openpyxl.load_workbook('학습.xlsx')
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 301):
            if sheet["A" + str(i)].value == "-":
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = learn[2]
                await client.send_message(message.channel, "★ 현재 사용중인 데이터 저장용량 : " + str(i) + "/300 ★")
                break
        file.save("학습.xlsx")

    if message.content.startswith("!기억"):
        file = openpyxl.load_workbook("학습.xlsx")
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == memory[1]:
                await client.send_message(message.channel, sheet["B" + str(i)].value)
                break

    if message.content.startswith("!역할설정"):
        role = ""
        rolename = message.content.split(" ")
        member = discord.utils.get(client.get_all_members(), id=rolename[1])
        for i in message.server.roles:
            if i.name == rolename[2]:
                role = i
                break
        await client.add_roles(member, role)


    if message.content.startswith("!경고"):
        memid = message.content.split(" ")
        file = openpyxl.load_workbook("경고.xlsx")
        sheet = file.active
        for i in range(1, 31):
            if str(sheet["A" + str(i)].value) == str(memid[1]):
                sheet["B" + str(i)].value = int(sheet["B" + str(i)].value) + 1
                break
            if sheet["B" + str(i)].value == 3:
                await client.add_roles(author, 뮤트)
                break
            if str(sheet["A" + str(i)].value) == "-":
                sheet["A" + str(i)].value = str(memid[1])
                sheet["B" + str(i)].value = 1
                break
        file.save("경고.xlsx")
        await client.send_message(message.channel, "경고를 받았디 주의하기 바란다 현재 누적경고 :" + str(sheet["B" + str(i)].value) + "회")

    if message.content.startswith("!확인"):
        memid = message.content.split(" ")
        file = openpyxl.load_workbook("경고.xlsx")
        sheet = file.active
        for i in range(1, 31):
            if str(sheet["A" + str(i)].value) == str(memid[1]):
                sheet["B" + str(i)].value = int(sheet["B" + str(i)].value) + 0
                break
        await client.send_message(message.channel, "누적 경고횟수는" + str(sheet["B" + str(i)].value) + "회")

    if message.content.startswith("!리셋"):
        memid = message.content.split(" ")
        file = openpyxl.load_workbook("경고.xlsx")
        sheet = file.active
        for i in range(1, 31):
            if str(sheet["A" + str(i)].value) == str(memid[1]):
                sheet["B" + str(i)].value = 0
                break
        file.save("경고.xlsx")
        await client.send_message(message.channel, "초기화 완료디 누적경고횟수는" + str(sheet["B" + str(i)].value) + "회")

    if message.content.startswith("!출석"):
        memid = message.author.id
        file = openpyxl.load_workbook("출석.xlsx")
        sheet = file.active
        for i in range(1, 31):
            if str(sheet["A" + str(i)].value) == str(memid[1]):
                sheet["B" + str(i)].value = int(sheet["B" + str(i)].value) + 1
                break
            if str(sheet["A" + str(i)].value) == "-":
                sheet["A" + str(i)].value = str(memid[1])
                sheet["B" + str(i)].value = 1
                break
        file.save("출석.xlsx")
        await client.send_message(message.channel, "출석이 확인되었디 " + memid + "의 현재 누적출석일수 :" + str(sheet["B" + str(i)].value) + "일")

    if message.content.startswith("!정산"):
        role = "명예린민"
        member = message.author.id
        await client.add_roles(member, "명예린민")
        await client.send_message(message.channel, "출석횟수가 2회 누적되었으므로 명예린민을 지급했디")



client.run("NDI0Nzc3MTEwOTE4Mzk3OTU1.XQUjQw.lQZSAtZQI-P_3Ffcf2QQ-cW6KXQ")
